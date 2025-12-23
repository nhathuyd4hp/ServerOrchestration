import os
import tempfile
from datetime import datetime

import pandas as pd
import redis
from celery import shared_task
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from src.core.config import settings
from src.core.logger import Log
from src.core.redis import REDIS_POOL
from src.robot.SeikyuNgoaiHanwa.api import APISharePoint
from src.robot.SeikyuNgoaiHanwa.automation import Excel, WebAccess


@shared_task(bind=True, name="Seikyu Ngoài Hanwa")
def seikyu(
    self,
    from_date: datetime | str,
    to_date: datetime | str,
):
    # ----- Type Check
    if isinstance(from_date, str):
        from_date = datetime.strptime(from_date, "%Y-%m-%d %H:%M:%S.%f").date()
    if isinstance(to_date, str):
        to_date = datetime.strptime(to_date, "%Y-%m-%d %H:%M:%S.%f").date()
    # ----- Logger
    logger = Log.get_logger(channel=self.request.id, redis_client=redis.Redis(connection_pool=REDIS_POOL))
    logger.info(f"{from_date} ~ {to_date}")
    with (
        tempfile.TemporaryDirectory() as temp_dir,
        sync_playwright() as p,
    ):
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        # --- #
        SeikyuFile: str = (
            "≪ベトナム≫請求書　阪和以外　(9日AM(10日分)、14日AM(15日分)、19日AM(20日分)、29日AM(末日分)に完成).xlsm"
        )
        if (
            APISharePoint(
                TENANT_ID=settings.API_SHAREPOINT_TENANT_ID,
                CLIENT_ID=settings.API_SHAREPOINT_CLIENT_ID,
                CLIENT_SECRET=settings.API_SHAREPOINT_CLIENT_SECRET,
            ).download_item(
                site_id="nskkogyo.sharepoint.com,fcec6ca2-58f4-4488-abf8-34e8ffbb741d,3136a8b2-a506-44d2-ad49-324bd156147c",
                breadcrumb=f"◆請求書　ベトナム専用◆/{SeikyuFile}",
                save_to=temp_dir,
            )
            is None
        ):
            logger.warning(f"Không tìm thấy file `{SeikyuFile}`")
            raise FileNotFoundError(f"Không tìm thấy file `{SeikyuFile}`")
        SeikyuFile = os.path.join(temp_dir, SeikyuFile)

        data: pd.DataFrame = WebAccess(
            username=settings.WEBACCESS_USERNAME,
            password=settings.WEBACCESS_PASSWORD,
            playwright=p,
            browser=browser,
            context=context,
        ).download_data(from_date, to_date)
        logger.info(f"Raw Data Shape: {data.shape}")
        if data.empty:
            return
        data = data[~data["商社名"].str.contains("阪和", na=False)]
        Excel.clear_contents(file_path=SeikyuFile, sheet_name="ACCESS貼り付け")
        Excel.write(file_path=SeikyuFile, data=data.columns.to_list(), sheet_name="ACCESS貼り付け")
        Excel.write(file_path=SeikyuFile, data=data.values, sheet_name="ACCESS貼り付け", cell_range="A2")
        Excel.macro(file_path=SeikyuFile, name="narabekae")
        rows, _ = pd.read_excel(
            io=SeikyuFile,
            sheet_name="請求一覧",
            header=None,
        ).shape
        # -- Clear Content 請求一覧 -- #
        for i in ["F", "H", "I"]:
            Excel.clear_contents(
                file_path=SeikyuFile,
                cell_range=f"{i}11:{i}{rows}",
                sheet_name="請求一覧",
                visible=False,
            )
        # Tìm dòng có công thức
        wb = load_workbook(SeikyuFile, data_only=False)
        ws = wb["請求一覧"]
        first_row_with_formula = None
        for cell in ws["A"]:
            row = cell.row
            if row < 11:
                continue
            if (
                all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 1])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 2])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 3])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 4])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 5])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 6])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 7])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 8])
                and all(c.value is None or (isinstance(c.value, str) and c.value.startswith("=")) for c in ws[row + 9])
            ):
                first_row_with_formula = row
                break
        # -- Copy Data -- #
        for sheet_name, sheet_data in Excel.read(
            file_path=SeikyuFile,
            visible=False,
        ):
            if sheet_name == "ACCESS貼り付け":
                sheet_data.columns = sheet_data.iloc[0]
                sheet_data = sheet_data.iloc[1:].reset_index(drop=True)
                Excel.write(
                    file_path=SeikyuFile,
                    data=[[item] for item in sheet_data["受注NO"].to_list()],
                    cell_range=f"I{first_row_with_formula}",
                    sheet_name="請求一覧",
                    visible=False,
                )
        for sheet_name, sheet_data in Excel.read(SeikyuFile, visible=False):
            if sheet_name == "請求一覧":
                data = sheet_data
                break
        # -- Extract Data--#
        # result = []
        # quote_url = []
        # prices = []
